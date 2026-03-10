from flask import Flask, render_template, request, Response, send_file
import cv2
import numpy as np
from datetime import datetime
import mysql.connector
import csv
from pyzbar.pyzbar import decode
from openpyxl import Workbook, load_workbook
import face_recognition
import os
import threading

app = Flask(__name__)

mydb = mysql.connector.connect(
    host='localhost',
    user='root',
    passwd='1234',
    port='3306',
    database='krisanth',
    auth_plugin='mysql_native_password'
)

# Set the path where uploaded photos will be saved
UPLOAD_FOLDER = 'faces'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

cursor = mydb.cursor()

today = datetime.now()
formatted_date = today.strftime('%Y-%m-%d %H:%M:%S')
current_time = today.strftime("%H:%M:%S")

path = 'faces'
images = []
classNames = []
registered_names = []

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        register_number = request.form['registerNumber']
        name = request.form['name']
        photo = request.files['photo']

        # Create the path to save the photo
        filename = f"{register_number}_{name}.{photo.filename.split('.')[-1]}"
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)

        # Save the photo to the specified path
        photo.save(save_path)

        # Perform any further processing or database insertion if needed

        return f"Registration successful! Photo saved as: {filename}"

    return render_template('register.html')

def gen_frames():
    def get_register_number_and_name(filename):
        filename = os.path.splitext(filename)[0]
        parts = filename.split('_')
        if len(parts) == 2:
            register_number = parts[0]
            name = parts[1]
            return register_number, name
        else:
            return None, None

    mylist = os.listdir(path)
    print(mylist)

    for cl in mylist:
        register_number, name = get_register_number_and_name(cl)
        if register_number is not None and name is not None:
            curimg = cv2.imread(f'{path}/{cl}')
            images.append(curimg)
            classNames.append(name)
            registered_names.append(register_number)
            print(f"Register Number: {register_number}, Name: {name}")

    def encodings(images):
        encodelist = []
        for img in images:
            img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            encode = face_recognition.face_encodings(img)[0]
            encodelist.append(encode)
        return encodelist

    encodelistKnown = encodings(images)
    print("Encoding Complete...")

    cap = cv2.VideoCapture(0)

    # Adjust the table and column names as per your database setup
    table_name = "Mark_Potrunga_Mam"
    column_register_number = "Register_Number"
    column_name = "Name"
    column_entry = "Entry"

    barcode_data = None  # Global variable to store the barcode data temporarily

    while True:
        success, img = cap.read()
        imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
        imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

        facesCurFrame = face_recognition.face_locations(imgS)
        encodeCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

        decoded_objects = decode(img)
        for obj in decoded_objects:
            barcode_data = obj.data.decode("utf-8")
            x, y, w, h = obj.rect
            cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
            print(barcode_data)

        for encodeFace, faceloc in zip(encodeCurFrame, facesCurFrame):
            matches = face_recognition.compare_faces(encodelistKnown, encodeFace)
            faceDis = face_recognition.face_distance(encodelistKnown, encodeFace)
            print(faceDis)
            matchIndex = np.argmin(faceDis)

            if matches[matchIndex]:
                name = classNames[matchIndex].upper()
                register_number = registered_names[matchIndex]
                print(f"Register Number: {register_number}, Name: {name}")
                y1, x2, y2, x1 = faceloc
                y1, x2, y2, x1 = y1 * 4, x2 * 4, y2 * 4, x1 * 4
                cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                cv2.rectangle(img, (x1, y2 - 35), (x2, y2), (0, 255, 0), cv2.FILLED)
                cv2.putText(img, name, (x1 + 6, y2 - 6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
                if name not in registered_names:
                    formatted_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    # Insert the new entry into the database
                    add_people = "INSERT INTO Mark_Potrunga_Mam (Register_Number, barcode, Name, Face_Time, Barcode_Time) VALUES (%s, %s, %s, %s, %s)"
                    students = (register_number, barcode_data, name, formatted_date, current_time)
                    cursor.execute(add_people, students)
                    mydb.commit()
                    # Add the name to the list of registered names
                    registered_names.append(name)
            else:
                print('error')

        ret, buffer = cv2.imencode('.jpg', img)
        if not ret:
            continue
        frame = buffer.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')


@app.route('/start_webcam')
def start_webcam():
    # Redirect to the webpage with the webcam streaming
    return render_template('start_webcam.html')

@app.route('/video_feed')
def video_feed():
    return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/download_attendance')
def download_attendance():
    # Fetch the attendance data from the database (assuming the table name is 'team_301')
    cursor.execute("SELECT Register_Number, Name, Face_Time, Barcode_Time,Status FROM Mark_Potrunga_Mam")
    attendance_data = cursor.fetchall()

    # Create a CSV file using the fetched data
    csv_filename = "attendance_list.csv"
    with open(csv_filename, 'w', newline='') as csvfile:
        csv_writer = csv.writer(csvfile)
        csv_writer.writerow(['Register Number', 'Name', 'Face_Time','Barcode_Time'])
        csv_writer.writerows(attendance_data)

    # Return the CSV file as a download response
    return send_file(csv_filename, as_attachment=True)


def update_attendance(barcode_data):
    # Load or create the Excel workbook
    try:
        workbook = load_workbook("attendance.xlsx")
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["ID", "Time"])

    # Get current date and time

    # Check if the barcode data is already in the attendance sheet
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
        if row[0] == barcode_data:
            # Update the out time
            worksheet.cell(row=worksheet.max_row, column=3, value=current_time)
            break
    else:
        # Add a new entry for the barcode data
        worksheet.append([barcode_data, current_time, ""])

    # Save the changes to the Excel file
    workbook.save("attendance.xlsx")


def gen_frames2():
    cap = cv2.VideoCapture(0)
    while True:
        success, frame = cap.read()  # read the camera frame
        if not success:
            print("Failed to capture frame")
            break
        else:
            ret, buffer = cv2.imencode('.jpg', frame)
            if not ret:
                print("Failed to encode frame")
                break

            decoded_objects = decode(frame)
            for obj in decoded_objects:
                barcode_data = obj.data.decode("utf-8")
                process_barcode_data(barcode_data)

            frame_bytes = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')  # concat frame one by one and show result



@app.route('/barcode')
def barcode():
    return render_template('barcode.html')


@app.route('/video_feed2')
def video_feed2():
    return Response(gen_frames2(), mimetype='multipart/x-mixed-replace; boundary=frame')


def process_barcode_data(barcode_data):
    print("Barcode Data:", barcode_data)
    update_attendance(barcode_data)


if __name__ == "__main__":
    app.run(debug=True)

