# import cv2
# from pyzbar.pyzbar import decode
# from openpyxl import Workbook, load_workbook
# from datetime import datetime
#
# def update_attendance(barcode_data):
#     # Load or create the Excel workbook
#     try:
#         workbook = load_workbook("attendance.xlsx")
#         worksheet = workbook.active
#     except FileNotFoundError:
#         workbook = Workbook()
#         worksheet = workbook.active
#         worksheet.append(["ID", "In Time", "Out Time"])
#
#     # Get current date and time
#     now = datetime.now()
#     current_time = now.strftime("%H:%M:%S")
#
#     # Check if the barcode data is already in the attendance sheet
#     for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
#         if row[0] == barcode_data:
#             # Update the out time
#             worksheet.cell(row=worksheet.max_row, column=3, value=current_time)
#             break
#     else:
#         # Add a new entry for the barcode data
#         worksheet.append([barcode_data, current_time, ""])
#
#     # Save the changes to the Excel file
#     workbook.save("attendance.xlsx")
#
# def scan_and_store_attendance():
#     # Initialize camera capture
#     cap = cv2.VideoCapture(0)
#
#     while True:
#         ret, frame = cap.read()
#
#         # Decode barcodes
#         decoded_objects = decode(frame)
#
#         for obj in decoded_objects:
#             # Draw a rectangle around the barcode
#             x, y, w, h = obj.rect
#             cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
#
#             # Print and store barcode data
#             barcode_data = obj.data.decode("utf-8")
#             print("Barcode Data:", barcode_data)
#             update_attendance(barcode_data)
#
#         # Display the frame
#         cv2.imshow('Smart Attendance System', frame)
#
#         # Check for key press to exit
#         if cv2.waitKey(1) & 0xFF == ord('q'):
#             break
#
#     # Release the camera and close all windows
#     cap.release()
#     cv2.destroyAllWindows()
#
# if __name__ == "__main__":
#     scan_and_store_attendance()


import cv2
from pyzbar.pyzbar import decode
from openpyxl import Workbook, load_workbook
from datetime import datetime
from flask import Flask, render_template, Response

app = Flask(__name__)


def update_attendance(barcode_data):
    # Load or create the Excel workbook
    try:
        workbook = load_workbook("attendance.xlsx")
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["ID", "In Time", "Out Time"])

    # Get current date and time
    now = datetime.now()
    current_time = now.strftime("%H:%M:%S")

    # Check if the barcode data is already in the attendance sheet
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, values_only=True):
        if row[0] == barcode_data:
            # Update the out time
            worksheet.cell(row=worksheet.max_row, column=3, value=current_time)
            break
    else:
        # Add a new entry for the barcode data
        worksheet.append([barcode_data, current_time, ""])

    # Save the changes to the Excel file
    workbook.save("attendance.xlsx")


def gen_frames():
    cap = cv2.VideoCapture(0)
    while True:
        success, frame = cap.read()  # read the camera frame
        if not success:
            print("Failed to capture frame")
            break
        else:
            ret, buffer = cv2.imencode('.jpg', frame)
            if not ret:
                print("Failed to encode frame")
                break
            frame_bytes = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')  # concat frame one by one and show result



@app.route('/')
def index():
    return render_template('barcode.html')


@app.route('/video_feed')
def video_feed():
    return Response(gen_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')


def process_barcode_data(barcode_data):
    print("Barcode Data:", barcode_data)
    update_attendance(barcode_data)


def decode_frames():
    cap = cv2.VideoCapture(0)
    while True:
        success, frame = cap.read()
        if not success:
            break

        decoded_objects = decode(frame)
        for obj in decoded_objects:
            barcode_data = obj.data.decode("utf-8")
            process_barcode_data(barcode_data)


if __name__ == "__main__":
    import threading

    threading.Thread(target=decode_frames, daemon=True).start()
    app.run(debug=True)
